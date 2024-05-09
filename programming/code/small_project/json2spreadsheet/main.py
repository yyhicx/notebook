# Copyright 2022 by chnxish. All Rights Reserved.
#
# Distributed under MIT license.
# See file LICENSE for detail or copy at https://opensource.org/licenses/MIT
"""Convert JSON data to Excel and generate an image."""
import json
import sys

import excel2img
import openpyxl

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python3 main.py json_file_name")
        sys.exit(1)

    json_file_name = sys.argv[1]
    with open(json_file_name, "r", encoding="utf-8") as f:
        json_object = json.load(f)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet"

    for row in json_object.get("data"):
        worksheet.append(row)

    for row in worksheet.rows:
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(
                horizontal="center", vertical="center"
            )

    for column_cells in worksheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells) + 5
        worksheet.column_dimensions[column_cells[0].column_letter].width = length

    thin_border = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style="thin"),
        right=openpyxl.styles.Side(style="thin"),
        top=openpyxl.styles.Side(style="thin"),
        bottom=openpyxl.styles.Side(style="thin"),
    )
    range_parameters = json_object.get("operation").get("thin_border")
    for row in worksheet.iter_rows(
        min_row=range_parameters.get("min_row"),
        max_row=range_parameters.get("max_row"),
        min_col=range_parameters.get("min_col"),
        max_col=range_parameters.get("max_col"),
    ):
        for cell in row:
            cell.border = thin_border

    for cell_range in json_object.get("operation").get("merge_cell"):
        worksheet.merge_cells(cell_range)

    ext_length = len(".json")
    excel_file_name = json_file_name[:-ext_length] + ".xlsx"
    image_file_name = json_file_name[:-ext_length] + ".png"
    workbook.save(excel_file_name)
    excel2img.export_img(excel_file_name, image_file_name, worksheet.title, None)
